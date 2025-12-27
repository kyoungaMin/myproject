"""
Create professional Excel spreadsheet for trending stocks data
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_trending_stocks_excel():
    """Create a professional Excel spreadsheet with trending stocks data"""

    # Create output directory if it doesn't exist
    output_dir = r"C:\myproject\output\data"
    os.makedirs(output_dir, exist_ok=True)

    # Create workbook and select active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "화제 종목"

    # Define data
    title = "화제 종목 TOP 5 | 2025.12.05"
    headers = ["순위", "티커", "종목명", "주가", "등락률"]
    data = [
        [1, "TSLA", "Tesla", "$385", "+8.7%"],
        [2, "NVDA", "NVIDIA", "$142", "+5.2%"],
        [3, "AAPL", "Apple", "$195", "+2.1%"]
    ]

    # Define styles
    title_font = Font(name='맑은 고딕', size=14, bold=True, color="1F4E78")
    header_font = Font(name='맑은 고딕', size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    data_font = Font(name='맑은 고딕', size=10)

    # Green fill for positive percentages
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(name='맑은 고딕', size=10, color="006100", bold=True)

    # Border style
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Title Row (Row 1)
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    # Header Row (Row 2)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    ws.row_dimensions[2].height = 20

    # Data Rows (Starting from Row 3)
    for row_num, row_data in enumerate(data, 3):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.font = data_font
            cell.border = thin_border

            # Alignment based on column
            if col_num == 1:  # 순위
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif col_num in [2, 3]:  # 티커, 종목명
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:  # 주가, 등락률
                cell.alignment = Alignment(horizontal='right', vertical='center')

            # Color coding for 등락률 column
            if col_num == 5:  # 등락률 column
                percentage_str = str(value)
                if percentage_str.startswith('+'):
                    cell.fill = green_fill
                    cell.font = green_font
                elif percentage_str.startswith('-'):
                    # Red fill for negative (if any)
                    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    red_font = Font(name='맑은 고딕', size=10, color="9C0006", bold=True)
                    cell.fill = red_fill
                    cell.font = red_font

    # Auto-fit column widths
    column_widths = {
        'A': 8,   # 순위
        'B': 10,  # 티커
        'C': 15,  # 종목명
        'D': 12,  # 주가
        'E': 12   # 등락률
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Save the workbook
    output_file = os.path.join(output_dir, "trending_2025-12-05.xlsx")
    wb.save(output_file)

    print("Excel file created successfully!")
    print(f"Location: {output_file}")
    print(f"Total rows: {len(data) + 2} (1 title + 1 header + {len(data)} data rows)")
    return output_file

if __name__ == "__main__":
    create_trending_stocks_excel()
